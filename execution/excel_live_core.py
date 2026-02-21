# execution/excel_live_core.py
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Any, Dict, Tuple, Optional

import openpyxl


CORE_VERSION = "2026-02-20.soft-volume-override.v1"


def _clamp(x: float, lo: float = 0.0, hi: float = 1.0) -> float:
    return max(lo, min(hi, x))


def _safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None:
            return default
        return float(v)
    except Exception:
        return default


def _parse_threshold_cell(s: Any) -> Optional[float]:
    """
    Accepts values like 'â‰¥0.60', '>=0.64', '0.50', etc.
    Returns float or None if not numeric.
    """
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    txt = str(s).strip()
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)", txt)
    if not m:
        return None
    return float(m.group(1))


def _env_float(name: str, default: float) -> float:
    v = os.getenv(name)
    if v is None or str(v).strip() == "":
        return default
    return _safe_float(v, default)


def _env_bool(name: str, default: bool = False) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    return str(v).strip().lower() in ("1", "true", "yes", "y", "on")


@dataclass
class CoreInputs:
    trend_strength: float          # 0..1
    structure_ok: bool             # True/False
    volume_score: float            # 0..1
    risk_state: str                # OK / REDUCE / KILL
    confidence_score: float        # 0..1
    volatility_regime: str         # LOW / NORMAL / EXTREME
    # optional macro flags (can be fed later)
    liquidity_regime: str = "EXPANSION"        # EXPANSION / CONTRACTION
    macro_risk_level: str = "LOW_RISK"         # LOW_RISK / HIGH_RISK
    shock_absorber: str = "NORMAL"             # NORMAL / REDUCE_EXPOSURE


class ExcelLiveCore:
    """
    Live Core evaluator based on workbook:
    - WEIGHT_THRESHOLD_MATRIX (weights + thresholds)

    FIX:
      If volume confirmation blocks too often (common on 15m),
      allow a SOFT override when AI score is very strong AND other gates pass.

      Default:
        vol_th = 0.50 (from Excel)
        soft relax = 0.10  => allow 0.40 when ai >= 0.75 and other gates ok
    """

    def __init__(self, workbook_path: str):
        if not os.path.exists(workbook_path):
            raise FileNotFoundError(f"EXCEL_MODEL_NOT_FOUND: {workbook_path}")

        self.wb = openpyxl.load_workbook(workbook_path, data_only=False)
        self.weights, self.thresholds = self._load_weight_threshold_matrix()

        # Soft override knobs (ENV)
        self.enable_soft_volume_override = _env_bool("ENABLE_SOFT_VOLUME_OVERRIDE", True)
        self.soft_volume_ai_min = _env_float("SOFT_VOLUME_AI_MIN", 0.75)
        self.soft_volume_relax = _env_float("SOFT_VOLUME_RELAX", 0.10)
        self.soft_volume_require_volband = _env_bool("SOFT_VOLUME_REQUIRE_VOLBAND", True)

    def _load_weight_threshold_matrix(self) -> Tuple[Dict[str, float], Dict[str, Any]]:
        ws = self.wb["WEIGHT_THRESHOLD_MATRIX"]

        weights: Dict[str, float] = {}
        thresholds: Dict[str, Any] = {}

        for r in range(2, ws.max_row + 1):
            comp = ws.cell(r, 1).value
            w = ws.cell(r, 2).value
            th = ws.cell(r, 3).value

            if not comp:
                continue

            comp_str = str(comp).strip().lower()
            weights[comp_str] = _safe_float(w, 0.0)
            thresholds[comp_str] = {"raw": th, "num": _parse_threshold_cell(th)}

        return weights, thresholds

    def _macro_gate(self, inp: CoreInputs) -> str:
        if inp.liquidity_regime == "CONTRACTION":
            return "BLOCK"
        if inp.macro_risk_level == "HIGH_RISK":
            return "BLOCK"
        if inp.shock_absorber == "REDUCE_EXPOSURE":
            return "BLOCK"
        return "ALLOW"

    def _vol_allowed(self, regime: str) -> bool:
        return regime in ("LOW", "NORMAL")

    def _score(self, inp: CoreInputs) -> float:
        w = self.weights

        trend_w = w.get("trend strength", 0.25)
        vol_w = w.get("volatility regime", 0.10)
        conf_w = w.get("confidence score", 0.15)
        risk_w = w.get("risk state modifier", 0.15)
        volconf_w = w.get("volume confirmation", 0.15)
        struct_w = w.get("structure validation", 0.20)

        risk_num = 1.0 if inp.risk_state == "OK" else (0.5 if inp.risk_state == "REDUCE" else 0.0)
        vol_num = 1.0 if inp.volatility_regime == "NORMAL" else (0.8 if inp.volatility_regime == "LOW" else 0.0)
        struct_num = 1.0 if inp.structure_ok else 0.0

        total = (
            inp.trend_strength * trend_w +
            struct_num * struct_w +
            inp.volume_score * volconf_w +
            risk_num * risk_w +
            inp.confidence_score * conf_w +
            vol_num * vol_w
        )
        return _clamp(total, 0.0, 1.0)

    def decide(self, inp: CoreInputs) -> Dict[str, Any]:
        ai_score = self._score(inp)
        macro_gate = self._macro_gate(inp)

        trend_th = (self.thresholds.get("trend strength", {}) or {}).get("num", 0.60) or 0.60
        vol_th = (self.thresholds.get("volume confirmation", {}) or {}).get("num", 0.50) or 0.50
        conf_th = (self.thresholds.get("confidence score", {}) or {}).get("num", 0.64) or 0.64

        trend_ok = inp.trend_strength >= float(trend_th)
        conf_ok = inp.confidence_score >= float(conf_th)
        struct_ok = bool(inp.structure_ok)
        risk_ok = inp.risk_state != "KILL"
        volband_ok = self._vol_allowed(inp.volatility_regime)

        # strict volume
        vol_ok_strict = inp.volume_score >= float(vol_th)

        # soft volume override
        soft_vol_th = _clamp(float(vol_th) - float(self.soft_volume_relax), 0.0, 1.0)
        vol_ok_soft = False

        if self.enable_soft_volume_override:
            other_gates_ok = (trend_ok and conf_ok and struct_ok and risk_ok)
            volband_req_ok = (volband_ok if self.soft_volume_require_volband else True)
            if other_gates_ok and volband_req_ok and ai_score >= float(self.soft_volume_ai_min):
                vol_ok_soft = inp.volume_score >= soft_vol_th

        vol_ok = vol_ok_strict or vol_ok_soft

        active_strategy = "YES" if (trend_ok and struct_ok and vol_ok and conf_ok and risk_ok and volband_ok) else "NO"
        final_trade_decision = "EXECUTE" if (macro_gate == "ALLOW" and active_strategy == "YES" and ai_score > 0.60) else "STAND_BY"

        return {
            "ai_score": ai_score,
            "macro_gate": macro_gate,
            "active_strategy": active_strategy,
            "final_trade_decision": final_trade_decision,
            "reasons": {
                "core_version": CORE_VERSION,

                "trend_strength": inp.trend_strength,
                "trend_th": float(trend_th),
                "trend_ok": trend_ok,

                "structure_ok": struct_ok,

                "volume_score": inp.volume_score,
                "volume_th": float(vol_th),
                "volume_th_soft": float(soft_vol_th),
                "volume_ok_strict": vol_ok_strict,
                "volume_ok_soft": vol_ok_soft,
                "volume_ok": vol_ok,

                "confidence_score": inp.confidence_score,
                "conf_th": float(conf_th),
                "confidence_ok": conf_ok,

                "risk_state": inp.risk_state,
                "risk_ok": risk_ok,

                "volatility_regime": inp.volatility_regime,
                "volband_ok": volband_ok,

                "soft_volume_override_enabled": bool(self.enable_soft_volume_override),
                "soft_volume_ai_min": float(self.soft_volume_ai_min),
                "soft_volume_relax": float(self.soft_volume_relax),
                "soft_volume_require_volband": bool(self.soft_volume_require_volband),
            },
        }
